import csv
import html
import json
import logging
import multiprocessing as mp
import random
import string
import time
from copy import deepcopy
from io import BytesIO, StringIO
from operator import itemgetter
import cherrypy
from cherrypy import _cperror
from mako.lookup import TemplateLookup
from mako.template import Template
import openpyxl
import secure
from sflib import SpiderFoot
from sfscan import startSpiderFootScanner
from spiderfoot import SpiderFootDb
from spiderfoot import SpiderFootHelpers
from spiderfoot import __version__
from spiderfoot.logger import logListenerSetup, logWorkerSetup
mp.set_start_method('spawn', force=True)

class SpiderFootWebUi:
    """SpiderFoot web interface."""
    lookup = TemplateLookup(directories=[''])
    defaultConfig = dict()
    config = dict()
    token = None
    docroot = ''

    def __init__(self: 'SpiderFootWebUi', web_config: dict, config: dict, loggingQueue: 'logging.handlers.QueueListener'=None) -> None:
        if False:
            i = 10
            return i + 15
        'Initialize web server.\n\n        Args:\n            web_config (dict): config settings for web interface (interface, port, root path)\n            config (dict): SpiderFoot config\n            loggingQueue: TBD\n\n        Raises:\n            TypeError: arg type is invalid\n            ValueError: arg value is invalid\n        '
        if not isinstance(config, dict):
            raise TypeError(f'config is {type(config)}; expected dict()')
        if not config:
            raise ValueError('config is empty')
        if not isinstance(web_config, dict):
            raise TypeError(f'web_config is {type(web_config)}; expected dict()')
        if not config:
            raise ValueError('web_config is empty')
        self.docroot = web_config.get('root', '/').rstrip('/')
        self.defaultConfig = deepcopy(config)
        dbh = SpiderFootDb(self.defaultConfig, init=True)
        sf = SpiderFoot(self.defaultConfig)
        self.config = sf.configUnserialize(dbh.configGet(), self.defaultConfig)
        if loggingQueue is None:
            self.loggingQueue = mp.Queue()
            logListenerSetup(self.loggingQueue, self.config)
        else:
            self.loggingQueue = loggingQueue
        logWorkerSetup(self.loggingQueue)
        self.log = logging.getLogger(f'spiderfoot.{__name__}')
        cherrypy.config.update({'error_page.401': self.error_page_401, 'error_page.404': self.error_page_404, 'request.error_response': self.error_page})
        csp = secure.ContentSecurityPolicy().default_src("'self'").script_src("'self'", "'unsafe-inline'", 'blob:').style_src("'self'", "'unsafe-inline'").base_uri("'self'").connect_src("'self'", 'data:').frame_src("'self'", 'data:').img_src("'self'", 'data:')
        secure_headers = secure.Secure(server=secure.Server().set('server'), cache=secure.CacheControl().must_revalidate(), csp=csp, referrer=secure.ReferrerPolicy().no_referrer())
        cherrypy.config.update({'tools.response_headers.on': True, 'tools.response_headers.headers': secure_headers.framework.cherrypy()})

    def error_page(self: 'SpiderFootWebUi') -> None:
        if False:
            return 10
        'Error page.'
        cherrypy.response.status = 500
        if self.config.get('_debug'):
            cherrypy.response.body = _cperror.get_error_page(status=500, traceback=_cperror.format_exc())
        else:
            cherrypy.response.body = b'<html><body>Error</body></html>'

    def error_page_401(self: 'SpiderFootWebUi', status: str, message: str, traceback: str, version: str) -> str:
        if False:
            return 10
        'Unauthorized access HTTP 401 error page.\n\n        Args:\n            status (str): HTTP response status code and message\n            message (str): Error message\n            traceback (str): Error stack trace\n            version (str): CherryPy version\n\n        Returns:\n            str: HTML response\n        '
        return ''

    def error_page_404(self: 'SpiderFootWebUi', status: str, message: str, traceback: str, version: str) -> str:
        if False:
            for i in range(10):
                print('nop')
        'Not found error page 404.\n\n        Args:\n            status (str): HTTP response status code and message\n            message (str): Error message\n            traceback (str): Error stack trace\n            version (str): CherryPy version\n\n        Returns:\n            str: HTTP response template\n        '
        templ = Template(filename='spiderfoot/templates/error.tmpl', lookup=self.lookup)
        return templ.render(message='Not Found', docroot=self.docroot, status=status, version=__version__)

    def jsonify_error(self: 'SpiderFootWebUi', status: str, message: str) -> dict:
        if False:
            while True:
                i = 10
        'Jsonify error response.\n\n        Args:\n            status (str): HTTP response status code and message\n            message (str): Error message\n\n        Returns:\n            dict: HTTP error response template\n        '
        cherrypy.response.headers['Content-Type'] = 'application/json'
        cherrypy.response.status = status
        return {'error': {'http_status': status, 'message': message}}

    def error(self: 'SpiderFootWebUi', message: str) -> None:
        if False:
            i = 10
            return i + 15
        'Show generic error page with error message.\n\n        Args:\n            message (str): error message\n\n        Returns:\n            None\n        '
        templ = Template(filename='spiderfoot/templates/error.tmpl', lookup=self.lookup)
        return templ.render(message=message, docroot=self.docroot, version=__version__)

    def cleanUserInput(self: 'SpiderFootWebUi', inputList: list) -> list:
        if False:
            i = 10
            return i + 15
        'Convert data to HTML entities; except quotes and ampersands.\n\n        Args:\n            inputList (list): list of strings to sanitize\n\n        Returns:\n            list: sanitized input\n\n        Raises:\n            TypeError: inputList type was invalid\n\n        Todo:\n            Review all uses of this function, then remove it.\n            Use of this function is overloaded.\n        '
        if not isinstance(inputList, list):
            raise TypeError(f'inputList is {type(inputList)}; expected list()')
        ret = list()
        for item in inputList:
            if not item:
                ret.append('')
                continue
            c = html.escape(item, True)
            c = c.replace('&amp;', '&').replace('&quot;', '"')
            ret.append(c)
        return ret

    def searchBase(self: 'SpiderFootWebUi', id: str=None, eventType: str=None, value: str=None) -> list:
        if False:
            while True:
                i = 10
        'Search.\n\n        Args:\n            id (str): scan ID\n            eventType (str): TBD\n            value (str): TBD\n\n        Returns:\n            list: search results\n        '
        retdata = []
        if not id and (not eventType) and (not value):
            return retdata
        if not value:
            value = ''
        regex = ''
        if value.startswith('/') and value.endswith('/'):
            regex = value[1:len(value) - 1]
            value = ''
        value = value.replace('*', '%')
        if value in [None, ''] and regex in [None, '']:
            value = '%'
            regex = ''
        dbh = SpiderFootDb(self.config)
        criteria = {'scan_id': id or '', 'type': eventType or '', 'value': value or '', 'regex': regex or ''}
        try:
            data = dbh.search(criteria)
        except Exception:
            return retdata
        for row in data:
            lastseen = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(row[0]))
            escapeddata = html.escape(row[1])
            escapedsrc = html.escape(row[2])
            retdata.append([lastseen, escapeddata, escapedsrc, row[3], row[5], row[6], row[7], row[8], row[10], row[11], row[4], row[13], row[14]])
        return retdata

    def buildExcel(self: 'SpiderFootWebUi', data: list, columnNames: list, sheetNameIndex: int=0) -> str:
        if False:
            print('Hello World!')
        'Convert supplied raw data into GEXF (Graph Exchange XML Format) format (e.g. for Gephi).\n\n        Args:\n            data (list): Scan result as list\n            columnNames (list): column names\n            sheetNameIndex (int): TBD\n\n        Returns:\n            str: Excel workbook\n        '
        rowNums = dict()
        workbook = openpyxl.Workbook()
        defaultSheet = workbook.active
        columnNames.pop(sheetNameIndex)
        allowed_sheet_chars = string.ascii_uppercase + string.digits + '_'
        for row in data:
            sheetName = ''.join([c for c in str(row.pop(sheetNameIndex)) if c.upper() in allowed_sheet_chars])
            try:
                sheet = workbook[sheetName]
            except KeyError:
                workbook.create_sheet(sheetName)
                sheet = workbook[sheetName]
                for (col_num, column_title) in enumerate(columnNames, 1):
                    cell = sheet.cell(row=1, column=col_num)
                    cell.value = column_title
                rowNums[sheetName] = 2
            for (col_num, cell_value) in enumerate(row, 1):
                cell = sheet.cell(row=rowNums[sheetName], column=col_num)
                cell.value = cell_value
            rowNums[sheetName] += 1
        if rowNums:
            workbook.remove(defaultSheet)
        workbook._sheets.sort(key=lambda ws: ws.title)
        with BytesIO() as f:
            workbook.save(f)
            f.seek(0)
            return f.read()

    @cherrypy.expose
    def scanexportlogs(self: 'SpiderFootWebUi', id: str, dialect: str='excel') -> bytes:
        if False:
            return 10
        'Get scan log\n\n        Args:\n            id (str): scan ID\n            dialect (str): CSV dialect (default: excel)\n\n        Returns:\n            bytes: scan logs in CSV format\n        '
        dbh = SpiderFootDb(self.config)
        try:
            data = dbh.scanLogs(id, None, None, True)
        except Exception:
            return self.error('Scan ID not found.')
        if not data:
            return self.error('Scan ID not found.')
        fileobj = StringIO()
        parser = csv.writer(fileobj, dialect=dialect)
        parser.writerow(['Date', 'Component', 'Type', 'Event', 'Event ID'])
        for row in data:
            parser.writerow([time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(row[0] / 1000)), str(row[1]), str(row[2]), str(row[3]), row[4]])
        cherrypy.response.headers['Content-Disposition'] = f'attachment; filename=SpiderFoot-{id}.log.csv'
        cherrypy.response.headers['Content-Type'] = 'application/csv'
        cherrypy.response.headers['Pragma'] = 'no-cache'
        return fileobj.getvalue().encode('utf-8')

    @cherrypy.expose
    def scancorrelationsexport(self: 'SpiderFootWebUi', id: str, filetype: str='csv', dialect: str='excel') -> str:
        if False:
            while True:
                i = 10
        'Get scan correlation data in CSV or Excel format.\n\n        Args:\n            id (str): scan ID\n            filetype (str): type of file ("xlsx|excel" or "csv")\n            dialect (str): CSV dialect (default: excel)\n\n        Returns:\n            str: results in CSV or Excel format\n        '
        dbh = SpiderFootDb(self.config)
        try:
            scaninfo = dbh.scanInstanceGet(id)
            scan_name = scaninfo[0]
        except Exception:
            return json.dumps(['ERROR', 'Could not retrieve info for scan.']).encode('utf-8')
        try:
            correlations = dbh.scanCorrelationList(id)
        except Exception:
            return json.dumps(['ERROR', 'Could not retrieve correlations for scan.']).encode('utf-8')
        headings = ['Rule Name', 'Correlation', 'Risk', 'Description']
        if filetype.lower() in ['xlsx', 'excel']:
            rows = []
            for row in correlations:
                correlation = row[1]
                rule_name = row[2]
                rule_risk = row[3]
                rule_description = row[5]
                rows.append([rule_name, correlation, rule_risk, rule_description])
            if scan_name:
                fname = f'{scan_name}-SpiderFoot-correlations.xlxs'
            else:
                fname = 'SpiderFoot-correlations.xlxs'
            cherrypy.response.headers['Content-Disposition'] = f'attachment; filename={fname}'
            cherrypy.response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            cherrypy.response.headers['Pragma'] = 'no-cache'
            return self.buildExcel(rows, headings, sheetNameIndex=0)
        if filetype.lower() == 'csv':
            fileobj = StringIO()
            parser = csv.writer(fileobj, dialect=dialect)
            parser.writerow(headings)
            for row in correlations:
                correlation = row[1]
                rule_name = row[2]
                rule_risk = row[3]
                rule_description = row[5]
                parser.writerow([rule_name, correlation, rule_risk, rule_description])
            if scan_name:
                fname = f'{scan_name}-SpiderFoot-correlations.csv'
            else:
                fname = 'SpiderFoot-correlations.csv'
            cherrypy.response.headers['Content-Disposition'] = f'attachment; filename={fname}'
            cherrypy.response.headers['Content-Type'] = 'application/csv'
            cherrypy.response.headers['Pragma'] = 'no-cache'
            return fileobj.getvalue().encode('utf-8')
        return self.error('Invalid export filetype.')

    @cherrypy.expose
    def scaneventresultexport(self: 'SpiderFootWebUi', id: str, type: str, filetype: str='csv', dialect: str='excel') -> str:
        if False:
            print('Hello World!')
        'Get scan event result data in CSV or Excel format\n\n        Args:\n            id (str): scan ID\n            type (str): TBD\n            filetype (str): type of file ("xlsx|excel" or "csv")\n            dialect (str): CSV dialect (default: excel)\n\n        Returns:\n            str: results in CSV or Excel format\n        '
        dbh = SpiderFootDb(self.config)
        data = dbh.scanResultEvent(id, type)
        if filetype.lower() in ['xlsx', 'excel']:
            rows = []
            for row in data:
                if row[4] == 'ROOT':
                    continue
                lastseen = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(row[0]))
                datafield = str(row[1]).replace('<SFURL>', '').replace('</SFURL>', '')
                rows.append([lastseen, str(row[4]), str(row[3]), str(row[2]), row[13], datafield])
            fname = 'SpiderFoot.xlsx'
            cherrypy.response.headers['Content-Disposition'] = f'attachment; filename={fname}'
            cherrypy.response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            cherrypy.response.headers['Pragma'] = 'no-cache'
            return self.buildExcel(rows, ['Updated', 'Type', 'Module', 'Source', 'F/P', 'Data'], sheetNameIndex=1)
        if filetype.lower() == 'csv':
            fileobj = StringIO()
            parser = csv.writer(fileobj, dialect=dialect)
            parser.writerow(['Updated', 'Type', 'Module', 'Source', 'F/P', 'Data'])
            for row in data:
                if row[4] == 'ROOT':
                    continue
                lastseen = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(row[0]))
                datafield = str(row[1]).replace('<SFURL>', '').replace('</SFURL>', '')
                parser.writerow([lastseen, str(row[4]), str(row[3]), str(row[2]), row[13], datafield])
            fname = 'SpiderFoot.csv'
            cherrypy.response.headers['Content-Disposition'] = f'attachment; filename={fname}'
            cherrypy.response.headers['Content-Type'] = 'application/csv'
            cherrypy.response.headers['Pragma'] = 'no-cache'
            return fileobj.getvalue().encode('utf-8')
        return self.error('Invalid export filetype.')

    @cherrypy.expose
    def scaneventresultexportmulti(self: 'SpiderFootWebUi', ids: str, filetype: str='csv', dialect: str='excel') -> str:
        if False:
            print('Hello World!')
        'Get scan event result data in CSV or Excel format for multiple scans\n\n        Args:\n            ids (str): comma separated list of scan IDs\n            filetype (str): type of file ("xlsx|excel" or "csv")\n            dialect (str): CSV dialect (default: excel)\n\n        Returns:\n            str: results in CSV or Excel format\n        '
        dbh = SpiderFootDb(self.config)
        scaninfo = dict()
        data = list()
        scan_name = ''
        for id in ids.split(','):
            scaninfo[id] = dbh.scanInstanceGet(id)
            if scaninfo[id] is None:
                continue
            scan_name = scaninfo[id][0]
            data = data + dbh.scanResultEvent(id)
        if not data:
            return None
        if filetype.lower() in ['xlsx', 'excel']:
            rows = []
            for row in data:
                if row[4] == 'ROOT':
                    continue
                lastseen = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(row[0]))
                datafield = str(row[1]).replace('<SFURL>', '').replace('</SFURL>', '')
                rows.append([scaninfo[row[12]][0], lastseen, str(row[4]), str(row[3]), str(row[2]), row[13], datafield])
            if len(ids.split(',')) > 1 or scan_name == '':
                fname = 'SpiderFoot.xlsx'
            else:
                fname = scan_name + '-SpiderFoot.xlsx'
            cherrypy.response.headers['Content-Disposition'] = f'attachment; filename={fname}'
            cherrypy.response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            cherrypy.response.headers['Pragma'] = 'no-cache'
            return self.buildExcel(rows, ['Scan Name', 'Updated', 'Type', 'Module', 'Source', 'F/P', 'Data'], sheetNameIndex=2)
        if filetype.lower() == 'csv':
            fileobj = StringIO()
            parser = csv.writer(fileobj, dialect=dialect)
            parser.writerow(['Scan Name', 'Updated', 'Type', 'Module', 'Source', 'F/P', 'Data'])
            for row in data:
                if row[4] == 'ROOT':
                    continue
                lastseen = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(row[0]))
                datafield = str(row[1]).replace('<SFURL>', '').replace('</SFURL>', '')
                parser.writerow([scaninfo[row[12]][0], lastseen, str(row[4]), str(row[3]), str(row[2]), row[13], datafield])
            if len(ids.split(',')) > 1 or scan_name == '':
                fname = 'SpiderFoot.csv'
            else:
                fname = scan_name + '-SpiderFoot.csv'
            cherrypy.response.headers['Content-Disposition'] = f'attachment; filename={fname}'
            cherrypy.response.headers['Content-Type'] = 'application/csv'
            cherrypy.response.headers['Pragma'] = 'no-cache'
            return fileobj.getvalue().encode('utf-8')
        return self.error('Invalid export filetype.')

    @cherrypy.expose
    def scansearchresultexport(self: 'SpiderFootWebUi', id: str, eventType: str=None, value: str=None, filetype: str='csv', dialect: str='excel') -> str:
        if False:
            for i in range(10):
                print('nop')
        'Get search result data in CSV or Excel format\n\n        Args:\n            id (str): scan ID\n            eventType (str): TBD\n            value (str): TBD\n            filetype (str): type of file ("xlsx|excel" or "csv")\n            dialect (str): CSV dialect (default: excel)\n\n        Returns:\n            str: results in CSV or Excel format\n        '
        data = self.searchBase(id, eventType, value)
        if not data:
            return None
        if filetype.lower() in ['xlsx', 'excel']:
            rows = []
            for row in data:
                if row[10] == 'ROOT':
                    continue
                datafield = str(row[1]).replace('<SFURL>', '').replace('</SFURL>', '')
                rows.append([row[0], str(row[10]), str(row[3]), str(row[2]), row[11], datafield])
            cherrypy.response.headers['Content-Disposition'] = 'attachment; filename=SpiderFoot.xlsx'
            cherrypy.response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            cherrypy.response.headers['Pragma'] = 'no-cache'
            return self.buildExcel(rows, ['Updated', 'Type', 'Module', 'Source', 'F/P', 'Data'], sheetNameIndex=1)
        if filetype.lower() == 'csv':
            fileobj = StringIO()
            parser = csv.writer(fileobj, dialect=dialect)
            parser.writerow(['Updated', 'Type', 'Module', 'Source', 'F/P', 'Data'])
            for row in data:
                if row[10] == 'ROOT':
                    continue
                datafield = str(row[1]).replace('<SFURL>', '').replace('</SFURL>', '')
                parser.writerow([row[0], str(row[10]), str(row[3]), str(row[2]), row[11], datafield])
            cherrypy.response.headers['Content-Disposition'] = 'attachment; filename=SpiderFoot.csv'
            cherrypy.response.headers['Content-Type'] = 'application/csv'
            cherrypy.response.headers['Pragma'] = 'no-cache'
            return fileobj.getvalue().encode('utf-8')
        return self.error('Invalid export filetype.')

    @cherrypy.expose
    def scanexportjsonmulti(self: 'SpiderFootWebUi', ids: str) -> str:
        if False:
            return 10
        'Get scan event result data in JSON format for multiple scans.\n\n        Args:\n            ids (str): comma separated list of scan IDs\n\n        Returns:\n            str: results in JSON format\n        '
        dbh = SpiderFootDb(self.config)
        scaninfo = list()
        scan_name = ''
        for id in ids.split(','):
            scan = dbh.scanInstanceGet(id)
            if scan is None:
                continue
            scan_name = scan[0]
            for row in dbh.scanResultEvent(id):
                lastseen = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(row[0]))
                event_data = str(row[1]).replace('<SFURL>', '').replace('</SFURL>', '')
                source_data = str(row[2])
                source_module = str(row[3])
                event_type = row[4]
                false_positive = row[13]
                if event_type == 'ROOT':
                    continue
                scaninfo.append({'data': event_data, 'event_type': event_type, 'module': source_module, 'source_data': source_data, 'false_positive': false_positive, 'last_seen': lastseen, 'scan_name': scan_name, 'scan_target': scan[1]})
        if len(ids.split(',')) > 1 or scan_name == '':
            fname = 'SpiderFoot.json'
        else:
            fname = scan_name + '-SpiderFoot.json'
        cherrypy.response.headers['Content-Disposition'] = f'attachment; filename={fname}'
        cherrypy.response.headers['Content-Type'] = 'application/json; charset=utf-8'
        cherrypy.response.headers['Pragma'] = 'no-cache'
        return json.dumps(scaninfo).encode('utf-8')

    @cherrypy.expose
    def scanviz(self: 'SpiderFootWebUi', id: str, gexf: str='0') -> str:
        if False:
            i = 10
            return i + 15
        'Export entities from scan results for visualising.\n\n        Args:\n            id (str): scan ID\n            gexf (str): TBD\n\n        Returns:\n            str: GEXF data\n        '
        if not id:
            return None
        dbh = SpiderFootDb(self.config)
        data = dbh.scanResultEvent(id, filterFp=True)
        scan = dbh.scanInstanceGet(id)
        if not scan:
            return None
        scan_name = scan[0]
        root = scan[1]
        if gexf == '0':
            return SpiderFootHelpers.buildGraphJson([root], data)
        if not scan_name:
            fname = 'SpiderFoot.gexf'
        else:
            fname = scan_name + 'SpiderFoot.gexf'
        cherrypy.response.headers['Content-Disposition'] = f'attachment; filename={fname}'
        cherrypy.response.headers['Content-Type'] = 'application/gexf'
        cherrypy.response.headers['Pragma'] = 'no-cache'
        return SpiderFootHelpers.buildGraphGexf([root], 'SpiderFoot Export', data)

    @cherrypy.expose
    def scanvizmulti(self: 'SpiderFootWebUi', ids: str, gexf: str='1') -> str:
        if False:
            return 10
        'Export entities results from multiple scans in GEXF format.\n\n        Args:\n            ids (str): scan IDs\n            gexf (str): TBD\n\n        Returns:\n            str: GEXF data\n        '
        dbh = SpiderFootDb(self.config)
        data = list()
        roots = list()
        scan_name = ''
        if not ids:
            return None
        for id in ids.split(','):
            scan = dbh.scanInstanceGet(id)
            if not scan:
                continue
            data = data + dbh.scanResultEvent(id, filterFp=True)
            roots.append(scan[1])
            scan_name = scan[0]
        if not data:
            return None
        if gexf == '0':
            return None
        if len(ids.split(',')) > 1 or scan_name == '':
            fname = 'SpiderFoot.gexf'
        else:
            fname = scan_name + '-SpiderFoot.gexf'
        cherrypy.response.headers['Content-Disposition'] = f'attachment; filename={fname}'
        cherrypy.response.headers['Content-Type'] = 'application/gexf'
        cherrypy.response.headers['Pragma'] = 'no-cache'
        return SpiderFootHelpers.buildGraphGexf(roots, 'SpiderFoot Export', data)

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanopts(self: 'SpiderFootWebUi', id: str) -> dict:
        if False:
            while True:
                i = 10
        'Return configuration used for the specified scan as JSON.\n\n        Args:\n            id: scan ID\n\n        Returns:\n            dict: scan options for the specified scan\n        '
        dbh = SpiderFootDb(self.config)
        ret = dict()
        meta = dbh.scanInstanceGet(id)
        if not meta:
            return ret
        if meta[3] != 0:
            started = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(meta[3]))
        else:
            started = 'Not yet'
        if meta[4] != 0:
            finished = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(meta[4]))
        else:
            finished = 'Not yet'
        ret['meta'] = [meta[0], meta[1], meta[2], started, finished, meta[5]]
        ret['config'] = dbh.scanConfigGet(id)
        ret['configdesc'] = dict()
        for key in list(ret['config'].keys()):
            if ':' not in key:
                globaloptdescs = self.config['__globaloptdescs__']
                if globaloptdescs:
                    ret['configdesc'][key] = globaloptdescs.get(key, f'{key} (legacy)')
            else:
                [modName, modOpt] = key.split(':')
                if modName not in list(self.config['__modules__'].keys()):
                    continue
                if modOpt not in list(self.config['__modules__'][modName]['optdescs'].keys()):
                    continue
                ret['configdesc'][key] = self.config['__modules__'][modName]['optdescs'][modOpt]
        return ret

    @cherrypy.expose
    def rerunscan(self: 'SpiderFootWebUi', id: str) -> None:
        if False:
            while True:
                i = 10
        'Rerun a scan.\n\n        Args:\n            id (str): scan ID\n\n        Returns:\n            None\n\n        Raises:\n            HTTPRedirect: redirect to info page for new scan\n        '
        cfg = deepcopy(self.config)
        modlist = list()
        dbh = SpiderFootDb(cfg)
        info = dbh.scanInstanceGet(id)
        if not info:
            return self.error('Invalid scan ID.')
        scanname = info[0]
        scantarget = info[1]
        scanconfig = dbh.scanConfigGet(id)
        if not scanconfig:
            return self.error(f'Error loading config from scan: {id}')
        modlist = scanconfig['_modulesenabled'].split(',')
        if 'sfp__stor_stdout' in modlist:
            modlist.remove('sfp__stor_stdout')
        targetType = SpiderFootHelpers.targetTypeFromString(scantarget)
        if not targetType:
            targetType = SpiderFootHelpers.targetTypeFromString(f'"{scantarget}"')
        if targetType not in ['HUMAN_NAME', 'BITCOIN_ADDRESS']:
            scantarget = scantarget.lower()
        scanId = SpiderFootHelpers.genScanInstanceId()
        try:
            p = mp.Process(target=startSpiderFootScanner, args=(self.loggingQueue, scanname, scanId, scantarget, targetType, modlist, cfg))
            p.daemon = True
            p.start()
        except Exception as e:
            self.log.error(f'[-] Scan [{scanId}] failed: {e}')
            return self.error(f'[-] Scan [{scanId}] failed: {e}')
        while dbh.scanInstanceGet(scanId) is None:
            self.log.info('Waiting for the scan to initialize...')
            time.sleep(1)
        raise cherrypy.HTTPRedirect(f'{self.docroot}/scaninfo?id={scanId}', status=302)

    @cherrypy.expose
    def rerunscanmulti(self: 'SpiderFootWebUi', ids: str) -> str:
        if False:
            print('Hello World!')
        'Rerun scans.\n\n        Args:\n            ids (str): comma separated list of scan IDs\n\n        Returns:\n            str: Scan list page HTML\n        '
        cfg = deepcopy(self.config)
        modlist = list()
        dbh = SpiderFootDb(cfg)
        for id in ids.split(','):
            info = dbh.scanInstanceGet(id)
            if not info:
                return self.error('Invalid scan ID.')
            scanconfig = dbh.scanConfigGet(id)
            scanname = info[0]
            scantarget = info[1]
            targetType = None
            if len(scanconfig) == 0:
                return self.error('Something went wrong internally.')
            modlist = scanconfig['_modulesenabled'].split(',')
            if 'sfp__stor_stdout' in modlist:
                modlist.remove('sfp__stor_stdout')
            targetType = SpiderFootHelpers.targetTypeFromString(scantarget)
            if targetType is None:
                return self.error('Invalid target type. Could not recognize it as a target SpiderFoot supports.')
            scanId = SpiderFootHelpers.genScanInstanceId()
            try:
                p = mp.Process(target=startSpiderFootScanner, args=(self.loggingQueue, scanname, scanId, scantarget, targetType, modlist, cfg))
                p.daemon = True
                p.start()
            except Exception as e:
                self.log.error(f'[-] Scan [{scanId}] failed: {e}')
                return self.error(f'[-] Scan [{scanId}] failed: {e}')
            while dbh.scanInstanceGet(scanId) is None:
                self.log.info('Waiting for the scan to initialize...')
                time.sleep(1)
        templ = Template(filename='spiderfoot/templates/scanlist.tmpl', lookup=self.lookup)
        return templ.render(rerunscans=True, docroot=self.docroot, pageid='SCANLIST', version=__version__)

    @cherrypy.expose
    def newscan(self: 'SpiderFootWebUi') -> str:
        if False:
            return 10
        'Configure a new scan.\n\n        Returns:\n            str: New scan page HTML\n        '
        dbh = SpiderFootDb(self.config)
        types = dbh.eventTypes()
        templ = Template(filename='spiderfoot/templates/newscan.tmpl', lookup=self.lookup)
        return templ.render(pageid='NEWSCAN', types=types, docroot=self.docroot, modules=self.config['__modules__'], scanname='', selectedmods='', scantarget='', version=__version__)

    @cherrypy.expose
    def clonescan(self: 'SpiderFootWebUi', id: str) -> str:
        if False:
            return 10
        'Clone an existing scan (pre-selected options in the newscan page).\n\n        Args:\n            id (str): scan ID to clone\n\n        Returns:\n            str: New scan page HTML pre-populated with options from cloned scan.\n        '
        dbh = SpiderFootDb(self.config)
        types = dbh.eventTypes()
        info = dbh.scanInstanceGet(id)
        if not info:
            return self.error('Invalid scan ID.')
        scanconfig = dbh.scanConfigGet(id)
        scanname = info[0]
        scantarget = info[1]
        targetType = None
        if scanname == '' or scantarget == '' or len(scanconfig) == 0:
            return self.error('Something went wrong internally.')
        targetType = SpiderFootHelpers.targetTypeFromString(scantarget)
        if targetType is None:
            scantarget = '&quot;' + scantarget + '&quot;'
        modlist = scanconfig['_modulesenabled'].split(',')
        templ = Template(filename='spiderfoot/templates/newscan.tmpl', lookup=self.lookup)
        return templ.render(pageid='NEWSCAN', types=types, docroot=self.docroot, modules=self.config['__modules__'], selectedmods=modlist, scanname=str(scanname), scantarget=str(scantarget), version=__version__)

    @cherrypy.expose
    def index(self: 'SpiderFootWebUi') -> str:
        if False:
            for i in range(10):
                print('nop')
        'Show scan list page.\n\n        Returns:\n            str: Scan list page HTML\n        '
        templ = Template(filename='spiderfoot/templates/scanlist.tmpl', lookup=self.lookup)
        return templ.render(pageid='SCANLIST', docroot=self.docroot, version=__version__)

    @cherrypy.expose
    def scaninfo(self: 'SpiderFootWebUi', id: str) -> str:
        if False:
            while True:
                i = 10
        'Information about a selected scan.\n\n        Args:\n            id (str): scan id\n\n        Returns:\n            str: scan info page HTML\n        '
        dbh = SpiderFootDb(self.config)
        res = dbh.scanInstanceGet(id)
        if res is None:
            return self.error('Scan ID not found.')
        templ = Template(filename='spiderfoot/templates/scaninfo.tmpl', lookup=self.lookup, input_encoding='utf-8')
        return templ.render(id=id, name=html.escape(res[0]), status=res[5], docroot=self.docroot, version=__version__, pageid='SCANLIST')

    @cherrypy.expose
    def opts(self: 'SpiderFootWebUi', updated: str=None) -> str:
        if False:
            return 10
        'Show module and global settings page.\n\n        Args:\n            updated (str): scan options were updated successfully\n\n        Returns:\n            str: scan options page HTML\n        '
        templ = Template(filename='spiderfoot/templates/opts.tmpl', lookup=self.lookup)
        self.token = random.SystemRandom().randint(0, 99999999)
        return templ.render(opts=self.config, pageid='SETTINGS', token=self.token, version=__version__, updated=updated, docroot=self.docroot)

    @cherrypy.expose
    def optsexport(self: 'SpiderFootWebUi', pattern: str=None) -> str:
        if False:
            while True:
                i = 10
        'Export configuration.\n\n        Args:\n            pattern (str): TBD\n\n        Returns:\n            str: Configuration settings\n        '
        sf = SpiderFoot(self.config)
        conf = sf.configSerialize(self.config)
        content = ''
        for opt in sorted(conf):
            if ':_' in opt or opt.startswith('_'):
                continue
            if pattern:
                if pattern in opt:
                    content += f'{opt}={conf[opt]}\n'
            else:
                content += f'{opt}={conf[opt]}\n'
        cherrypy.response.headers['Content-Disposition'] = 'attachment; filename="SpiderFoot.cfg"'
        cherrypy.response.headers['Content-Type'] = 'text/plain'
        return content

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def optsraw(self: 'SpiderFootWebUi') -> str:
        if False:
            while True:
                i = 10
        'Return global and module settings as json.\n\n        Returns:\n            str: settings as JSON\n        '
        ret = dict()
        self.token = random.SystemRandom().randint(0, 99999999)
        for opt in self.config:
            if not opt.startswith('__'):
                ret['global.' + opt] = self.config[opt]
                continue
            if opt == '__modules__':
                for mod in sorted(self.config['__modules__'].keys()):
                    for mo in sorted(self.config['__modules__'][mod]['opts'].keys()):
                        if mo.startswith('_'):
                            continue
                        ret['module.' + mod + '.' + mo] = self.config['__modules__'][mod]['opts'][mo]
        return ['SUCCESS', {'token': self.token, 'data': ret}]

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scandelete(self: 'SpiderFootWebUi', id: str) -> str:
        if False:
            i = 10
            return i + 15
        'Delete scan(s).\n\n        Args:\n            id (str): comma separated list of scan IDs\n\n        Returns:\n            str: JSON response\n        '
        if not id:
            return self.jsonify_error('404', 'No scan specified')
        dbh = SpiderFootDb(self.config)
        ids = id.split(',')
        for scan_id in ids:
            res = dbh.scanInstanceGet(scan_id)
            if not res:
                return self.jsonify_error('404', f'Scan {scan_id} does not exist')
            if res[5] in ['RUNNING', 'STARTING', 'STARTED']:
                return self.jsonify_error('400', f'Scan {scan_id} is {res[5]}. You cannot delete running scans.')
        for scan_id in ids:
            dbh.scanInstanceDelete(scan_id)
        return ''

    @cherrypy.expose
    def savesettings(self: 'SpiderFootWebUi', allopts: str, token: str, configFile: 'cherrypy._cpreqbody.Part'=None) -> None:
        if False:
            i = 10
            return i + 15
        'Save settings, also used to completely reset them to default.\n\n        Args:\n            allopts: TBD\n            token (str): CSRF token\n            configFile (cherrypy._cpreqbody.Part): TBD\n\n        Returns:\n            None\n\n        Raises:\n            HTTPRedirect: redirect to scan settings\n        '
        if str(token) != str(self.token):
            return self.error(f'Invalid token ({token})')
        if configFile and configFile.file:
            try:
                contents = configFile.file.read()
                if isinstance(contents, bytes):
                    contents = contents.decode('utf-8')
                tmp = dict()
                for line in contents.split('\n'):
                    if '=' not in line:
                        continue
                    opt_array = line.strip().split('=')
                    if len(opt_array) == 1:
                        opt_array[1] = ''
                    tmp[opt_array[0]] = '='.join(opt_array[1:])
                allopts = json.dumps(tmp).encode('utf-8')
            except Exception as e:
                return self.error(f'Failed to parse input file. Was it generated from SpiderFoot? ({e})')
        if allopts == 'RESET':
            if self.reset_settings():
                raise cherrypy.HTTPRedirect(f'{self.docroot}/opts?updated=1')
            return self.error('Failed to reset settings')
        try:
            dbh = SpiderFootDb(self.config)
            useropts = json.loads(allopts)
            cleanopts = dict()
            for opt in list(useropts.keys()):
                cleanopts[opt] = self.cleanUserInput([useropts[opt]])[0]
            currentopts = deepcopy(self.config)
            sf = SpiderFoot(self.config)
            self.config = sf.configUnserialize(cleanopts, currentopts)
            dbh.configSet(sf.configSerialize(self.config))
        except Exception as e:
            return self.error(f'Processing one or more of your inputs failed: {e}')
        raise cherrypy.HTTPRedirect(f'{self.docroot}/opts?updated=1')

    @cherrypy.expose
    def savesettingsraw(self: 'SpiderFootWebUi', allopts: str, token: str) -> str:
        if False:
            i = 10
            return i + 15
        'Save settings, also used to completely reset them to default.\n\n        Args:\n            allopts: TBD\n            token (str): CSRF token\n\n        Returns:\n            str: save success as JSON\n        '
        cherrypy.response.headers['Content-Type'] = 'application/json; charset=utf-8'
        if str(token) != str(self.token):
            return json.dumps(['ERROR', f'Invalid token ({token}).']).encode('utf-8')
        if allopts == 'RESET':
            if self.reset_settings():
                return json.dumps(['SUCCESS', '']).encode('utf-8')
            return json.dumps(['ERROR', 'Failed to reset settings']).encode('utf-8')
        try:
            dbh = SpiderFootDb(self.config)
            useropts = json.loads(allopts)
            cleanopts = dict()
            for opt in list(useropts.keys()):
                cleanopts[opt] = self.cleanUserInput([useropts[opt]])[0]
            currentopts = deepcopy(self.config)
            sf = SpiderFoot(self.config)
            self.config = sf.configUnserialize(cleanopts, currentopts)
            dbh.configSet(sf.configSerialize(self.config))
        except Exception as e:
            return json.dumps(['ERROR', f'Processing one or more of your inputs failed: {e}']).encode('utf-8')
        return json.dumps(['SUCCESS', '']).encode('utf-8')

    def reset_settings(self: 'SpiderFootWebUi') -> bool:
        if False:
            return 10
        'Reset settings to default.\n\n        Returns:\n            bool: success\n        '
        try:
            dbh = SpiderFootDb(self.config)
            dbh.configClear()
            self.config = deepcopy(self.defaultConfig)
        except Exception:
            return False
        return True

    @cherrypy.expose
    def resultsetfp(self: 'SpiderFootWebUi', id: str, resultids: str, fp: str) -> str:
        if False:
            for i in range(10):
                print('nop')
        'Set a bunch of results (hashes) as false positive.\n\n        Args:\n            id (str): scan ID\n            resultids (str): comma separated list of result IDs\n            fp (str): 0 or 1\n\n        Returns:\n            str: set false positive status as JSON\n        '
        cherrypy.response.headers['Content-Type'] = 'application/json; charset=utf-8'
        dbh = SpiderFootDb(self.config)
        if fp not in ['0', '1']:
            return json.dumps(['ERROR', 'No FP flag set or not set correctly.']).encode('utf-8')
        try:
            ids = json.loads(resultids)
        except Exception:
            return json.dumps(['ERROR', 'No IDs supplied.']).encode('utf-8')
        status = dbh.scanInstanceGet(id)
        if not status:
            return self.error(f'Invalid scan ID: {id}')
        if status[5] not in ['ABORTED', 'FINISHED', 'ERROR-FAILED']:
            return json.dumps(['WARNING', 'Scan must be in a finished state when setting False Positives.']).encode('utf-8')
        if fp == '0':
            data = dbh.scanElementSourcesDirect(id, ids)
            for row in data:
                if str(row[14]) == '1':
                    return json.dumps(['WARNING', f'Cannot unset element {id} as False Positive if a parent element is still False Positive.']).encode('utf-8')
        childs = dbh.scanElementChildrenAll(id, ids)
        allIds = ids + childs
        ret = dbh.scanResultsUpdateFP(id, allIds, fp)
        if ret:
            return json.dumps(['SUCCESS', '']).encode('utf-8')
        return json.dumps(['ERROR', 'Exception encountered.']).encode('utf-8')

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def eventtypes(self: 'SpiderFootWebUi') -> list:
        if False:
            return 10
        'List all event types.\n\n        Returns:\n            list: list of event types\n        '
        cherrypy.response.headers['Content-Type'] = 'application/json; charset=utf-8'
        dbh = SpiderFootDb(self.config)
        types = dbh.eventTypes()
        ret = list()
        for r in types:
            ret.append([r[1], r[0]])
        return sorted(ret, key=itemgetter(0))

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def modules(self: 'SpiderFootWebUi') -> list:
        if False:
            for i in range(10):
                print('nop')
        'List all modules.\n\n        Returns:\n            list: list of modules\n        '
        cherrypy.response.headers['Content-Type'] = 'application/json; charset=utf-8'
        ret = list()
        modinfo = list(self.config['__modules__'].keys())
        if not modinfo:
            return ret
        modinfo.sort()
        for m in modinfo:
            if '__' in m:
                continue
            ret.append({'name': m, 'descr': self.config['__modules__'][m]['descr']})
        return ret

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def correlationrules(self: 'SpiderFootWebUi') -> list:
        if False:
            while True:
                i = 10
        'List all correlation rules.\n\n        Returns:\n            list: list of correlation rules\n        '
        cherrypy.response.headers['Content-Type'] = 'application/json; charset=utf-8'
        ret = list()
        rules = self.config['__correlationrules__']
        if not rules:
            return ret
        for r in rules:
            ret.append({'id': r['id'], 'name': r['meta']['name'], 'descr': r['meta']['description'], 'risk': r['meta']['risk']})
        return ret

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def ping(self: 'SpiderFootWebUi') -> list:
        if False:
            while True:
                i = 10
        'For the CLI to test connectivity to this server.\n\n        Returns:\n            list: SpiderFoot version as JSON\n        '
        return ['SUCCESS', __version__]

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def query(self: 'SpiderFootWebUi', query: str) -> str:
        if False:
            return 10
        'For the CLI to run queries against the database.\n\n        Args:\n            query (str): SQL query\n\n        Returns:\n            str: query results as JSON\n        '
        dbh = SpiderFootDb(self.config)
        if not query:
            return self.jsonify_error('400', 'Invalid query.')
        if not query.lower().startswith('select'):
            return self.jsonify_error('400', 'Non-SELECTs are unpredictable and not recommended.')
        try:
            ret = dbh.dbh.execute(query)
            data = ret.fetchall()
            columnNames = [c[0] for c in dbh.dbh.description]
            return [dict(zip(columnNames, row)) for row in data]
        except Exception as e:
            return self.jsonify_error('500', str(e))

    @cherrypy.expose
    def startscan(self: 'SpiderFootWebUi', scanname: str, scantarget: str, modulelist: str, typelist: str, usecase: str) -> str:
        if False:
            return 10
        'Initiate a scan.\n\n        Args:\n            scanname (str): scan name\n            scantarget (str): scan target\n            modulelist (str): comma separated list of modules to use\n            typelist (str): selected modules based on produced event data types\n            usecase (str): selected module group (passive, investigate, footprint, all)\n\n        Returns:\n            str: start scan status as JSON\n\n        Raises:\n            HTTPRedirect: redirect to new scan info page\n        '
        scanname = self.cleanUserInput([scanname])[0]
        scantarget = self.cleanUserInput([scantarget])[0]
        if not scanname:
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = 'application/json; charset=utf-8'
                return json.dumps(['ERROR', 'Incorrect usage: scan name was not specified.']).encode('utf-8')
            return self.error('Invalid request: scan name was not specified.')
        if not scantarget:
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = 'application/json; charset=utf-8'
                return json.dumps(['ERROR', 'Incorrect usage: scan target was not specified.']).encode('utf-8')
            return self.error('Invalid request: scan target was not specified.')
        if not typelist and (not modulelist) and (not usecase):
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = 'application/json; charset=utf-8'
                return json.dumps(['ERROR', 'Incorrect usage: no modules specified for scan.']).encode('utf-8')
            return self.error('Invalid request: no modules specified for scan.')
        targetType = SpiderFootHelpers.targetTypeFromString(scantarget)
        if targetType is None:
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = 'application/json; charset=utf-8'
                return json.dumps(['ERROR', 'Unrecognised target type.']).encode('utf-8')
            return self.error('Invalid target type. Could not recognize it as a target SpiderFoot supports.')
        dbh = SpiderFootDb(self.config)
        cfg = deepcopy(self.config)
        sf = SpiderFoot(cfg)
        modlist = list()
        if modulelist:
            modlist = modulelist.replace('module_', '').split(',')
        if len(modlist) == 0 and typelist:
            typesx = typelist.replace('type_', '').split(',')
            modlist = sf.modulesProducing(typesx)
            newmods = deepcopy(modlist)
            newmodcpy = deepcopy(newmods)
            while len(newmodcpy) > 0:
                for etype in sf.eventsToModules(newmodcpy):
                    xmods = sf.modulesProducing([etype])
                    for mod in xmods:
                        if mod not in modlist:
                            modlist.append(mod)
                            newmods.append(mod)
                newmodcpy = deepcopy(newmods)
                newmods = list()
        if len(modlist) == 0 and usecase:
            for mod in self.config['__modules__']:
                if usecase == 'all' or usecase in self.config['__modules__'][mod]['group']:
                    modlist.append(mod)
        if not modlist:
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = 'application/json; charset=utf-8'
                return json.dumps(['ERROR', 'Incorrect usage: no modules specified for scan.']).encode('utf-8')
            return self.error('Invalid request: no modules specified for scan.')
        if 'sfp__stor_db' not in modlist:
            modlist.append('sfp__stor_db')
        modlist.sort()
        if 'sfp__stor_stdout' in modlist:
            modlist.remove('sfp__stor_stdout')
        if targetType in ['HUMAN_NAME', 'USERNAME', 'BITCOIN_ADDRESS']:
            scantarget = scantarget.replace('"', '')
        else:
            scantarget = scantarget.lower()
        scanId = SpiderFootHelpers.genScanInstanceId()
        try:
            p = mp.Process(target=startSpiderFootScanner, args=(self.loggingQueue, scanname, scanId, scantarget, targetType, modlist, cfg))
            p.daemon = True
            p.start()
        except Exception as e:
            self.log.error(f'[-] Scan [{scanId}] failed: {e}')
            return self.error(f'[-] Scan [{scanId}] failed: {e}')
        while dbh.scanInstanceGet(scanId) is None:
            self.log.info('Waiting for the scan to initialize...')
            time.sleep(1)
        if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
            cherrypy.response.headers['Content-Type'] = 'application/json; charset=utf-8'
            return json.dumps(['SUCCESS', scanId]).encode('utf-8')
        raise cherrypy.HTTPRedirect(f'{self.docroot}/scaninfo?id={scanId}')

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def stopscan(self: 'SpiderFootWebUi', id: str) -> str:
        if False:
            while True:
                i = 10
        'Stop a scan.\n\n        Args:\n            id (str): comma separated list of scan IDs\n\n        Returns:\n            str: JSON response\n        '
        if not id:
            return self.jsonify_error('404', 'No scan specified')
        dbh = SpiderFootDb(self.config)
        ids = id.split(',')
        for scan_id in ids:
            res = dbh.scanInstanceGet(scan_id)
            if not res:
                return self.jsonify_error('404', f'Scan {scan_id} does not exist')
            scan_status = res[5]
            if scan_status == 'FINISHED':
                return self.jsonify_error('400', f'Scan {scan_id} has already finished.')
            if scan_status == 'ABORTED':
                return self.jsonify_error('400', f'Scan {scan_id} has already aborted.')
            if scan_status != 'RUNNING' and scan_status != 'STARTING':
                return self.jsonify_error('400', f"The running scan is currently in the state '{scan_status}', please try again later or restart SpiderFoot.")
        for scan_id in ids:
            dbh.scanInstanceSet(scan_id, status='ABORT-REQUESTED')
        return ''

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def vacuum(self):
        if False:
            for i in range(10):
                print('nop')
        dbh = SpiderFootDb(self.config)
        try:
            if dbh.vacuumDB():
                return json.dumps(['SUCCESS', '']).encode('utf-8')
            return json.dumps(['ERROR', 'Vacuuming the database failed']).encode('utf-8')
        except Exception as e:
            return json.dumps(['ERROR', f'Vacuuming the database failed: {e}']).encode('utf-8')

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanlog(self: 'SpiderFootWebUi', id: str, limit: str=None, rowId: str=None, reverse: str=None) -> list:
        if False:
            while True:
                i = 10
        'Scan log data.\n\n        Args:\n            id (str): scan ID\n            limit (str): TBD\n            rowId (str): TBD\n            reverse (str): TBD\n\n        Returns:\n            list: scan log\n        '
        dbh = SpiderFootDb(self.config)
        retdata = []
        try:
            data = dbh.scanLogs(id, limit, rowId, reverse)
        except Exception:
            return retdata
        for row in data:
            generated = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(row[0] / 1000))
            retdata.append([generated, row[1], row[2], html.escape(row[3]), row[4]])
        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanerrors(self: 'SpiderFootWebUi', id: str, limit: str=None) -> list:
        if False:
            while True:
                i = 10
        'Scan error data.\n\n        Args:\n            id (str): scan ID\n            limit (str): limit number of results\n\n        Returns:\n            list: scan errors\n        '
        dbh = SpiderFootDb(self.config)
        retdata = []
        try:
            data = dbh.scanErrors(id, limit)
        except Exception:
            return retdata
        for row in data:
            generated = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(row[0] / 1000))
            retdata.append([generated, row[1], html.escape(str(row[2]))])
        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanlist(self: 'SpiderFootWebUi') -> list:
        if False:
            print('Hello World!')
        'Produce a list of scans.\n\n        Returns:\n            list: scan list\n        '
        dbh = SpiderFootDb(self.config)
        data = dbh.scanInstanceList()
        retdata = []
        for row in data:
            created = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(row[3]))
            riskmatrix = {'HIGH': 0, 'MEDIUM': 0, 'LOW': 0, 'INFO': 0}
            correlations = dbh.scanCorrelationSummary(row[0], by='risk')
            if correlations:
                for c in correlations:
                    riskmatrix[c[0]] = c[1]
            if row[4] == 0:
                started = 'Not yet'
            else:
                started = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(row[4]))
            if row[5] == 0:
                finished = 'Not yet'
            else:
                finished = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(row[5]))
            retdata.append([row[0], row[1], row[2], created, started, finished, row[6], row[7], riskmatrix])
        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanstatus(self: 'SpiderFootWebUi', id: str) -> list:
        if False:
            i = 10
            return i + 15
        'Show basic information about a scan, including status and number of each event type.\n\n        Args:\n            id (str): scan ID\n\n        Returns:\n            list: scan status\n        '
        dbh = SpiderFootDb(self.config)
        data = dbh.scanInstanceGet(id)
        if not data:
            return []
        created = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(data[2]))
        started = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(data[3]))
        ended = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(data[4]))
        riskmatrix = {'HIGH': 0, 'MEDIUM': 0, 'LOW': 0, 'INFO': 0}
        correlations = dbh.scanCorrelationSummary(id, by='risk')
        if correlations:
            for c in correlations:
                riskmatrix[c[0]] = c[1]
        return [data[0], data[1], created, started, ended, data[5], riskmatrix]

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scansummary(self: 'SpiderFootWebUi', id: str, by: str) -> list:
        if False:
            print('Hello World!')
        'Summary of scan results.\n\n        Args:\n            id (str): scan ID\n            by (str): filter by type\n\n        Returns:\n            list: scan summary\n        '
        retdata = []
        dbh = SpiderFootDb(self.config)
        try:
            scandata = dbh.scanResultSummary(id, by)
        except Exception:
            return retdata
        try:
            statusdata = dbh.scanInstanceGet(id)
        except Exception:
            return retdata
        for row in scandata:
            if row[0] == 'ROOT':
                continue
            lastseen = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(row[2]))
            retdata.append([row[0], row[1], lastseen, row[3], row[4], statusdata[5]])
        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scancorrelations(self: 'SpiderFootWebUi', id: str) -> list:
        if False:
            return 10
        'Correlation results from a scan.\n\n        Args:\n            id (str): scan ID\n\n        Returns:\n            list: correlation result list\n        '
        retdata = []
        dbh = SpiderFootDb(self.config)
        try:
            corrdata = dbh.scanCorrelationList(id)
        except Exception:
            return retdata
        for row in corrdata:
            retdata.append([row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7]])
        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scaneventresults(self: 'SpiderFootWebUi', id: str, eventType: str=None, filterfp: bool=False, correlationId: str=None) -> list:
        if False:
            for i in range(10):
                print('nop')
        'Return all event results for a scan as JSON.\n\n        Args:\n            id (str): scan ID\n            eventType (str): filter by event type\n            filterfp (bool): remove false positives from search results\n            correlationId (str): filter by events associated with a correlation\n\n        Returns:\n            list: scan results\n        '
        retdata = []
        dbh = SpiderFootDb(self.config)
        if not eventType:
            eventType = 'ALL'
        try:
            data = dbh.scanResultEvent(id, eventType, filterfp, correlationId=correlationId)
        except Exception:
            return retdata
        for row in data:
            lastseen = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(row[0]))
            retdata.append([lastseen, html.escape(row[1]), html.escape(row[2]), row[3], row[5], row[6], row[7], row[8], row[13], row[14], row[4]])
        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scaneventresultsunique(self: 'SpiderFootWebUi', id: str, eventType: str, filterfp: bool=False) -> list:
        if False:
            for i in range(10):
                print('nop')
        'Return unique event results for a scan as JSON.\n\n        Args:\n            id (str): filter search results by scan ID\n            eventType (str): filter search results by event type\n            filterfp (bool): remove false positives from search results\n\n        Returns:\n            list: unique search results\n        '
        dbh = SpiderFootDb(self.config)
        retdata = []
        try:
            data = dbh.scanResultEventUnique(id, eventType, filterfp)
        except Exception:
            return retdata
        for row in data:
            escaped = html.escape(row[0])
            retdata.append([escaped, row[1], row[2]])
        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def search(self: 'SpiderFootWebUi', id: str=None, eventType: str=None, value: str=None) -> list:
        if False:
            while True:
                i = 10
        'Search scans.\n\n        Args:\n            id (str): filter search results by scan ID\n            eventType (str): filter search results by event type\n            value (str): filter search results by event value\n\n        Returns:\n            list: search results\n        '
        try:
            return self.searchBase(id, eventType, value)
        except Exception:
            return []

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanhistory(self: 'SpiderFootWebUi', id: str) -> list:
        if False:
            i = 10
            return i + 15
        'Historical data for a scan.\n\n        Args:\n            id (str): scan ID\n\n        Returns:\n            list: scan history\n        '
        if not id:
            return self.jsonify_error('404', 'No scan specified')
        dbh = SpiderFootDb(self.config)
        try:
            return dbh.scanResultHistory(id)
        except Exception:
            return []

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanelementtypediscovery(self: 'SpiderFootWebUi', id: str, eventType: str) -> dict:
        if False:
            for i in range(10):
                print('nop')
        'Scan element type discovery.\n\n        Args:\n            id (str): scan ID\n            eventType (str): filter by event type\n\n        Returns:\n            dict\n        '
        dbh = SpiderFootDb(self.config)
        pc = dict()
        datamap = dict()
        retdata = dict()
        try:
            leafSet = dbh.scanResultEvent(id, eventType)
            [datamap, pc] = dbh.scanElementSourcesAll(id, leafSet)
        except Exception:
            return retdata
        del pc['ROOT']
        retdata['tree'] = SpiderFootHelpers.dataParentChildToTree(pc)
        retdata['data'] = datamap
        return retdata